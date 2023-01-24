import json
import xlrd
import datetime
import openpyxl

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

from django.http import HttpResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User, Group
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.db.models import Q

from main.decorators import role_required, ajax_required, permissions_required
from main.functions import generate_form_errors, get_auto_id
from dishes.models import *
from main.functions import decrypt_message, encrypt_message, get_auto_id, get_otp
from dishes.forms import *
from shops.forms import FileForm
from general.models import UserActivity


class CuisineAutoComplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return Cuisine.objects.none()

        items = Cuisine.objects.filter(is_deleted=False)

        if self.q:
            items = items.filter(Q(auto_id__istartswith=self.q) | 
                                 Q(name__istartswith=self.q) 
                                )

        return items


class DishCategoryAutoComplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return DishCategory.objects.none()

        items = DishCategory.objects.filter(is_deleted=False)

        if self.q:
            items = items.filter(Q(auto_id__istartswith=self.q) | 
                                 Q(name__istartswith=self.q) 
                                )

        return items


@login_required
@permissions_required(['can_manage_cuisine'])
def cuisines(request):

    instances = Cuisine.objects.filter(is_deleted=False).order_by("-date_added")
    
    filter_data = {}
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(name__icontains=query) |
            Q(type__icontains=query)
        )
        filter_data['q'] = query
    
    context = {
        'instances': instances, 
        'page_name' : 'Cuisines', 
        'app_name' :'Dishes', 
        'page_title' : 'Cuisines', 
        'filter_data' : filter_data,
        'is_cuisine' : True,
    }
    
    return render(request, 'dashboard/dishes/cuisines.html', context)


@login_required
@permissions_required(['can_view_cuisine'])
def cuisine(request, pk):
    instance = Cuisine.objects.get(pk=pk, is_deleted=False)
    
    context = {
        'instance': instance, 
        'page_name' : 'Cuisine', 
        'app_name' :'Cuisines', 
        'page_title' : 'Cuisine', 
        'is_need_light_box' : True, 
        'is_cuisine' : True,
    }
         
    return render(request, 'dashboard/dishes/cuisine.html', context)


@login_required
@permissions_required(['can_create_cuisine'])
def cuisine_create(request):
    if request.method == "POST":
        form = CuisineForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.creator = request.user
            data.updater = request.user
            data.auto_id = get_auto_id(Cuisine)
            data.save()
            
            activity = UserActivity(
                user=request.user,
                activity_type="Create",
                app="Dishes",
                instance = data.name,
                title = "Created an Cuisine",
            )
            activity.save()

            response_data = {
                "status" : "success", 
                "stable" : "false", 
                "title" : "Successfully Created", 
                "message" : "Cuisine Successfully Created.", 
                "redirect" : "true", 
                
                "redirect_url" : reverse('dishes:cuisine', kwargs={"pk": data.pk})
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status" : "false", 
                "stable" : "true", 
                "title" : "Form validation error", 
                "message" : message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        form = CuisineForm()
        context = {
            "form" : form, 
            'app_name' :'Cuisine', 
            "page_title" : "Create Cuisine", 
            "redirect" : True, 
            "url" : reverse('dishes:cuisine_create'), 
            'is_cuisine' : True,
        }

        return render(request, 'dashboard/dishes/cuisine_entry.html', context)
    

@login_required
@permissions_required(['can_modify_cuisine'])
def cuisine_edit(request, pk):
    if request.method == "POST":
        instance = get_object_or_404(Cuisine.objects.filter(pk=pk))
        form = CuisineForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            data = form.save(commit=False)
            data.date_updated = datetime.datetime.now()
            data.updater = request.user
            data.save()
            
            activity = UserActivity(
                user=request.user,
                activity_type="Update",
                app="Dishes",
                instance = data.name,
                title = "Updated an Cuisine",
            )
            activity.save()

            response_data = {
                "status" : "success", 
                "stable" : "false", 
                "title" : "Successfully Updated", 
                "message" : "Cuisine Successfully Updated.", 
                "redirect" : "true", 
                "redirect_url" : reverse('dishes:cuisine', kwargs={"pk": data.pk})
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status" : "false", 
                "stable" : "true", 
                "title" : "Form validation error", 
                "message" : message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        instance = get_object_or_404(Cuisine.objects.filter(pk=pk))
        form = CuisineForm(instance=instance)
        context = {
            "form" : form, 
            'app_name' :'Cuisine', 
            "page_title" : "Edit Cuisine", 
            "redirect" : True, 
            "url" : reverse('dishes:cuisine_edit', kwargs={"pk":instance.pk}), 
            "is_edit" :True, 
            "instance" : instance, 
            'is_cuisine' : True,
        }

        return render(request, 'dashboard/dishes/cuisine_entry.html', context)


@login_required
@permissions_required(['can_delete_cuisine'])
@ajax_required
def cuisine_delete(request, pk):
    instance = Cuisine.objects.get(pk=pk)
    
    activity = UserActivity(
        user=request.user,
        activity_type="Delete",
        app="Dishes",
        instance = instance.name,
        title = "Deleted an Cuisine",
    )
    activity.save()
    
    Cuisine.objects.filter(pk=pk).update(is_deleted=True)

    response_data = {
        "status" : "success", 
        "title" : "Cuisine Deleted", 
        "message" : "Cuisine Successfully Deleted.", 
        "redirect" : "true", 
        "redirect_url" : reverse('dishes:cuisines')
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
@permissions_required(['can_manage_dish'])
def dishes(request):
    instances = Dish.objects.filter(is_deleted=False).order_by("-date_added")
    
    filter_data = {}
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(name__icontains=query) |
            Q(cuisine__name__icontains=query) |
            Q(dish_category__name__icontains=query) |
            Q(dish_timing__icontains=query) |
            Q(dietary_type__icontains=query) 
        )
        filter_data['q'] = query
    
    context = {
        'instances': instances, 
        'page_name' : 'Dishes', 
        'app_name' :'Dishes', 
        'page_title' : 'Dishes', 
        'filter_data' : filter_data,
        'is_dish' : True,
    }
    
    return render(request, 'dashboard/dishes/dishes.html', context)


@login_required
@permissions_required(['can_view_dish'])
def dish(request, pk):
    instance = Dish.objects.get(pk=pk, is_deleted=False)
    
    context = {
        'instance': instance, 
        'page_name' : 'Dish ', 
        'app_name' :'Dishes', 
        'page_title' : 'Dish ', 
        'is_need_light_box' : True, 
        'is_dish' : True,
    }
         
    return render(request, 'dashboard/dishes/dish.html', context)


@login_required
@permissions_required(['can_create_dish'])
def dish_create(request):
    if request.method == "POST":
        form = DishForm(request.POST, request.FILES)
            
        if form.is_valid():
            data = form.save(commit=False)
            data.creator = request.user
            data.updater = request.user
            
            if not  data.image:
                is_verified = False
            else:
                is_verified = True
                
            data.is_verified = is_verified
            data.auto_id = get_auto_id(Dish)
            data.save()
            
            activity = UserActivity(
                user=request.user,
                activity_type="Create",
                app="Dish",
                instance = data.name,
                title = "Created an Dish",
            )
            activity.save()

            response_data = {
                "status" : "success", 
                "stable" : "false", 
                "title" : "Successfully Created", 
                "message" : "Dish  Successfully Created.", 
                "redirect" : "true", 
                
                "redirect_url" : reverse('dishes:dish', kwargs={"pk": data.pk})
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status" : "false", 
                "stable" : "true", 
                "title" : "Form validation error", 
                "message" : message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        form = DishForm()
        context = {
            "form" : form, 
            'app_name' :'Dishes', 
            "page_title" : "Create Dish", 
            "redirect" : True, 
            "url" : reverse('dishes:dish_create'), 
            'is_dish' : True,
        }

        return render(request, 'dashboard/dishes/dish_entry.html', context)
    

@login_required
@permissions_required(['can_modify_dish'])
def dish_edit(request, pk):
    if request.method == "POST":
        instance = get_object_or_404(Dish.objects.filter(pk=pk))
        form = DishForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            data = form.save(commit=False)
            data.date_updated = datetime.datetime.now()
            data.updater = request.user
            
            if not data.image :
                is_verified = False
            else:
                is_verified = True
                
            data.is_verified = is_verified
            data.save()
            
            activity = UserActivity(
                user=request.user,
                activity_type="Update",
                app="Dish",
                instance = data.name,
                title = "Updated an Dish",
            )
            activity.save()

            response_data = {
                "status" : "success", 
                "stable" : "false", 
                "title" : "Successfully Updated", 
                "message" : "Dish Successfully Updated.", 
                "redirect" : "true", 
                "redirect_url" : reverse('dishes:dish', kwargs={"pk": data.pk})
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status" : "false", 
                "stable" : "true", 
                "title" : "Form validation error", 
                "message" : message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        instance = get_object_or_404(Dish.objects.filter(pk=pk))
        form = DishForm(instance=instance)
        context = {
            "form" : form, 
            'app_name' :'Dishes', 
            "page_title" : "Edit Dish", 
            "redirect" : True, 
            "url" : reverse('dishes:dish_edit', kwargs={"pk":instance.pk}), 
            "is_edit" :True, 
            "instance" : instance, 
            'is_dish' : True,
        }

        return render(request, 'dashboard/dishes/dish_entry.html', context)


@login_required
@permissions_required(['can_delete_dish'])
@ajax_required
def dish_delete(request, pk):
    instance = Dish.objects.get(pk=pk)
    activity = UserActivity(
        user=request.user,
        activity_type="Delete",
        app="Dish",
        instance = instance.name,
        title = "Deleted an Dish",
    )
    activity.save()
    
    Dish.objects.filter(pk=pk).update(is_deleted=True)

    response_data = {
        "status" : "success", 
        "title" : "Dish Deleted", 
        "message" : "Dish Successfully Deleted.", 
        "redirect" : "true", 
        "redirect_url" : reverse('dishes:dishes')
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
@permissions_required(['can_manage_dish_category'])
def dish_categories(request):
    instances = DishCategory.objects.filter(is_deleted=False).order_by("-date_added")
    
    filter_data = {}
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(name__icontains=query) 
        )
        filter_data['q'] = query
    
    context = {
        'instances': instances, 
        'page_name' : 'Categories', 
        'app_name' :'Dishes', 
        'page_title' : 'Categories', 
        'filter_data' : filter_data,
        'is_dc' : True,
    }
    
    return render(request, 'dashboard/dishes/categories.html', context)


@login_required
@permissions_required(['can_view_dish_category'])
def dish_category(request, pk):
    instance = DishCategory.objects.get(pk=pk, is_deleted=False)
    
    context = {
        'instance': instance, 
        'page_name' : 'Dish Category', 
        'app_name' :'Dishes', 
        'page_title' : 'Dish Category', 
        'is_need_light_box' : True, 
        'is_dc' : True,
    }
         
    return render(request, 'dashboard/dishes/category.html', context)


@login_required
@permissions_required(['can_create_dish_category'])
def dish_category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.save(commit=False)
            data.creator = request.user
            data.updater = request.user
            data.auto_id = get_auto_id(DishCategory)
            data.save()
            
            activity = UserActivity(
                user=request.user,
                activity_type="Create",
                app="Dish",
                instance = data.name,
                title = "Created an Dish Category",
            )
            activity.save()

            response_data = {
                "status" : "success", 
                "stable" : "false", 
                "title" : "Successfully Created", 
                "message" : "Dish category Successfully Created.", 
                "redirect" : "true", 
                
                "redirect_url" : reverse('dishes:dish_category', kwargs={"pk": data.pk})
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status" : "false", 
                "stable" : "true", 
                "title" : "Form validation error", 
                "message" : message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        form = CategoryForm()
        context = {
            "form" : form, 
            'app_name' :'Dishes', 
            "page_title" : "Create Category", 
            "redirect" : True, 
            "url" : reverse('dishes:dish_category_create'), 
            'is_dc' : True,
        }

        return render(request, 'dashboard/dishes/category_entry.html', context)
    

@login_required
@permissions_required(['can_modify_dish_category'])
def dish_category_edit(request, pk):
    if request.method == "POST":
        instance = get_object_or_404(DishCategory.objects.filter(pk=pk))
        form = CategoryForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            data = form.save(commit=False)
            data.date_updated = datetime.datetime.now()
            data.updater = request.user
            data.save()
            
            activity = UserActivity(
                user=request.user,
                activity_type="Update",
                app="Dish",
                instance = data.name,
                title = "Updated an Dish Category",
            )
            activity.save()

            response_data = {
                "status" : "success", 
                "stable" : "false", 
                "title" : "Successfully Updated", 
                "message" : "Dish Category Successfully Updated.", 
                "redirect" : "true", 
                "redirect_url" : reverse('dishes:dish_category', kwargs={"pk": data.pk})
            }
        else:
            message = generate_form_errors(form, formset=False)

            response_data = {
                "status" : "false", 
                "stable" : "true", 
                "title" : "Form validation error", 
                "message" : message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    else:
        instance = get_object_or_404(DishCategory.objects.filter(pk=pk))
        form = CategoryForm(instance=instance)
        context = {
            "form" : form, 
            'app_name' :'Dishes', 
            "page_title" : "Edit category", 
            "redirect" : True, 
            "url" : reverse('dishes:dish_category_edit', kwargs={"pk":instance.pk}), 
            "is_edit" :True, 
            "instance" : instance, 
            'is_dc' : True,
        }

        return render(request, 'dashboard/dishes/category_entry.html', context)


@login_required
@permissions_required(['can_delete_dish_category'])
@ajax_required
def dish_category_delete(request, pk):
    instance = DishCategory.objects.get(pk=pk)
    
    activity = UserActivity(
        user=request.user,
        activity_type="Delete",
        app="Dish",
        instance = instance.name,
        title = "Deleted an Dish Category",
    )
    activity.save()
    DishCategory.objects.filter(pk=pk).update(is_deleted=True)

    response_data = {
        "status" : "success", 
        "title" : "Dish Category Deleted", 
        "message" : "Dish Category Successfully Deleted.", 
        "redirect" : "true", 
        "redirect_url" : reverse('dishes:dish_categories')
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
@permissions_required(['can_upload_shops'])
def upload_dishes(request):
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        
        is_verified = False

        if form.is_valid():
            input_excel = request.FILES['file']
            book = xlrd.open_workbook(file_contents=input_excel.read())
            sheet = book.sheet_by_index(0)

            dict_list = []
            keys = [str(sheet.cell(0, col_index).value) for col_index in range(sheet.ncols)]
            for row_index in range(1, sheet.nrows):
                d = {keys[col_index]: str(sheet.cell(row_index, col_index).value)
                    for col_index in range(sheet.ncols)}
                dict_list.append(d)

            is_ok = True
            message = ''
            row_count = 0
            last_row = 0
            for item in dict_list:
                name = item['Name']
                cuisine = item['Cuisine']
                dish_category = item['Dish Category']
                dish_timing = item['Dish Timing']
                dietary_type = item['Dietary Type']
                row_count+=1
                last_row = row_count- 1
                
                if not Cuisine.objects.filter(name=cuisine).exists():
                    response_data = {
                        "status": "false", 
                        "stable": "true", 
                        "title": "Cuisine Not Found", 
                        "message": f"You need too create cuisine first in row : {row_count} , Data uploaded upto row number {last_row}"
                    }
                        
                    return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                else:
                    cuisine = Cuisine.objects.get(name=cuisine, is_deleted=False)
                
                if not DishCategory.objects.filter(name=dish_category).exists():
                    response_data = {
                        "status": "false", 
                        "stable": "true", 
                        "title": "Dish Category Not Found", 
                        "message": f"You need too create category first in row : {row_count} , Data uploaded upto row number {last_row}"
                    }
                        
                    return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                
                else:
                    dish_category = DishCategory.objects.get(name=dish_category, is_deleted=False)
                
                
                if  Dish.objects.filter(name=name, is_deleted=False).exists():
                    response_data =  {
                            "status": "false", 
                            "stable": "true", 
                            "title": "Dish Name Duplicates", 
                            "message": f"You need too check dish  in row : {row_count} , Data uploaded upto row number {last_row}"
                        }
                    
                    return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                else:
                    
                    dish = Dish.objects.create(
                       name = name, 
                       cuisine =cuisine, 
                       dish_category = dish_category, 
                       dish_timing = dish_timing, 
                       dietary_type = dietary_type, 
                       is_verified =is_verified, 
                       auto_id =  get_auto_id(Dish), 
                       date_added = datetime.datetime.now(), 
                       creator= request.user, 
                       updater=request.user, 
                    )
                    dish.save()
                    
                    activity = UserActivity(
                        user=request.user,
                        activity_type="Upload",
                        app="Dish",
                        instance = dish.name,
                        title = "Dish Uploaded",
                    )
                    activity.save()
                
                    
                    response_data = {
                        "status" : "success", 
                        "stable" : "false ", 
                        "message" : "Dishes Successfully Uploaded.", 
                        "redirect" : "true", 
                        "redirect_url" : reverse('dishes:dishes')
                    }

            return HttpResponse(json.dumps(response_data), content_type='application/javascript')
        else:
            form = FileForm()
            title = "Upload Restaurant"
            context = {
                "form" : form, 
                "title" : title, 

                "is_need_popup_box" : True, 
                "is_need_dropzone" : True
            }
            
            return render(request, 'dashboard/dishes/upload_dishes.html', context)
    else:
        form = FileForm()

        context = {
            "form" : form, 
            'page_name' : 'Upload Dishes ', 
            'app_name' :'Dishes', 
            'page_title' : 'Upload Dishes', 
            "redirect" : True, 
            "url" : reverse('dishes:upload_dishes'), 

            "is_need_popup_box" : True, 
            "is_need_dropzone" : True,
            'is_dish' : True,
        }
        
        return render(request, 'dashboard/dishes/upload_dishes.html', context)
    